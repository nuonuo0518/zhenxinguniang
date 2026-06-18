"""
verify_skin_rows.py
用途：写入完成后，读回 PlayerSkin.xlsx 验证每处改动是否已正确体现到 Excel 中
      逐行逐列对比期望值与实际值，输出完整改动清单及验证结果

用法：
    python verify_skin_rows.py --data skin_data.json

输入：与 write_skin_rows.py 相同格式的 JSON 文件
输出：每个 sheet 的每一行每一列的期望值 vs 实际值对比，以及整体 PASS/FAIL 汇总

依赖：openpyxl（只读，不修改文件）
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: 缺少 openpyxl，请先安装：pip install openpyxl")
    sys.exit(1)

FILE_PATH = r"F:\OL2wc\NBA2KOL2Doc_proj\PlayerSkin.xlsx"


def load_workbook(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: 文件不存在：{filepath}")
        sys.exit(1)
    return openpyxl.load_workbook(filepath, read_only=True, data_only=True)


def find_row_by_id(ws, target_id):
    """在 sheet 中按 ID 列（第1列）查找行，返回该行的所有值（tuple），找不到返回 None"""
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            if int(row[0]) == int(target_id):
                return row
        except (ValueError, TypeError):
            continue
    return None


def normalize(val):
    """统一化值类型用于比较：float 整数转 int，None 保持"""
    if val is None:
        return None
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


def verify_sheet(ws, rows_data, sheet_name):  # noqa: ARG001 - sheet_name used by caller for reporting
    """
    验证 sheet 中每一行的期望字段值是否与实际值一致
    返回 (passed, results) 其中 results 是逐行逐列的详情列表
    """
    all_passed = True
    results = []

    for row_data in rows_data:
        target_id = row_data.get("1") or row_data.get(1)
        if target_id is None:
            results.append({
                "id": "?",
                "found": False,
                "fields": [],
                "error": "JSON 中未找到列1(ID)字段"
            })
            all_passed = False
            continue

        actual_row = find_row_by_id(ws, target_id)
        if actual_row is None:
            results.append({
                "id": target_id,
                "found": False,
                "fields": [],
                "error": f"ID={target_id} 在 sheet 中不存在"
            })
            all_passed = False
            continue

        row_result = {"id": target_id, "found": True, "fields": []}

        for col_key, expected_val in row_data.items():
            col = int(col_key)
            # Excel 列从1开始，row tuple 从0开始
            actual_val = actual_row[col - 1] if col - 1 < len(actual_row) else None
            expected_norm = normalize(expected_val)
            actual_norm = normalize(actual_val)

            passed = (expected_norm == actual_norm)
            if not passed:
                all_passed = False

            row_result["fields"].append({
                "col": col,
                "expected": expected_val,
                "actual": actual_val,
                "passed": passed
            })

        results.append(row_result)

    return all_passed, results


def format_col_label(col):
    """将列号转为 Excel 列字母（1=A, 2=B...）"""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def print_report(sheet_name, verify_results):
    """打印单个 sheet 的完整验证报告"""
    print(f"\n{'=' * 60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'=' * 60}")

    for row_result in verify_results:
        target_id = row_result["id"]

        if not row_result["found"]:
            print(f"  [FAIL] ID={target_id}: {row_result.get('error', '未找到')}")
            continue

        row_pass = all(f["passed"] for f in row_result["fields"])
        status = "[OK]" if row_pass else "[FAIL]"
        print(f"\n  {status} ID={target_id} 行验证{'通过' if row_pass else '失败'}:")

        for field in row_result["fields"]:
            col = field["col"]
            col_label = format_col_label(col)
            expected = field["expected"]
            actual = field["actual"]
            passed = field["passed"]
            mark = "[OK]" if passed else "[FAIL]"

            if passed:
                print(f"    {mark} 列{col}({col_label}): {repr(expected)}")
            else:
                print(f"    {mark} 列{col}({col_label}): 期望={repr(expected)}  实际={repr(actual)}")


def main():
    parser = argparse.ArgumentParser(description="验证 PlayerSkin.xlsx 写入结果")
    parser.add_argument("--data", "-d", required=True, help="JSON 数据文件路径（与 write_skin_rows.py 相同格式）")
    parser.add_argument("--excel-dir", default=None, help="Excel 文件目录（默认使用内置路径）")
    args = parser.parse_args()

    if not os.path.exists(args.data):
        print(f"ERROR: 数据文件不存在：{args.data}")
        sys.exit(1)

    with open(args.data, "r", encoding="utf-8") as f:
        skin_data = json.load(f)

    filepath = FILE_PATH
    if args.excel_dir:
        filepath = os.path.join(args.excel_dir, "PlayerSkin.xlsx")

    print(f"=== PlayerSkin 写入验证 ===")
    print(f"数据文件：{args.data}")
    print(f"Excel：{filepath}")

    wb = load_workbook(filepath)
    total_sheets = 0
    passed_sheets = 0
    all_changes = []  # 汇总所有改动

    # 按依赖顺序验证
    write_order = ["UniformPanel", "PicPanel", "VideoPanel", "CelebratePanel", "IntroPanel", "Main", "UserInfoPanel"]

    for sheet_name in write_order:
        if sheet_name not in skin_data:
            continue

        rows = skin_data[sheet_name]
        if not isinstance(rows, list):
            rows = [rows]

        total_sheets += 1

        if sheet_name not in wb.sheetnames:
            print(f"\n[FAIL] Sheet '{sheet_name}' 不存在于 Excel 文件中")
            continue

        ws = wb[sheet_name]
        sheet_passed, results = verify_sheet(ws, rows, sheet_name)
        if sheet_passed:
            passed_sheets += 1

        print_report(sheet_name, results)

        # 收集改动清单（用于最终汇总）
        for row_result in results:
            for field in row_result.get("fields", []):
                all_changes.append({
                    "sheet": sheet_name,
                    "id": row_result["id"],
                    "col": field["col"],
                    "col_label": format_col_label(field["col"]),
                    "value": field["expected"],
                    "passed": field["passed"],
                    "actual": field["actual"],
                })

    wb.close()

    # 打印改动清单汇总
    print(f"\n{'=' * 60}")
    print(f"改动清单汇总（共 {len(all_changes)} 个字段）")
    print(f"{'=' * 60}")

    current_sheet = None
    current_id = None
    for change in all_changes:
        if change["sheet"] != current_sheet:
            current_sheet = change["sheet"]
            current_id = None
            print(f"\n[{current_sheet}]")
        if change["id"] != current_id:
            current_id = change["id"]
            print(f"  ID={current_id}:")
        mark = "[OK]" if change["passed"] else "[FAIL]"
        print(f"    {mark} 列{change['col']}({change['col_label']}): {repr(change['value'])}", end="")
        if not change["passed"]:
            print(f"  ← 实际值: {repr(change['actual'])}", end="")
        print()

    # 最终结论
    print(f"\n{'=' * 60}")
    failed_count = sum(1 for c in all_changes if not c["passed"])
    if failed_count == 0:
        print(f"[OK] 验证通过：全部 {len(all_changes)} 个字段已正确写入 Excel")
    else:
        print(f"[FAIL] 验证失败：{failed_count} 个字段不符，{len(all_changes) - failed_count} 个字段正确")
        sys.exit(1)


if __name__ == "__main__":
    main()
