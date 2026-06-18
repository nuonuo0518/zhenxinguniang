"""
read_sheet_info.py
用途：读取 PlayerSkin.xlsx 中各 sheet 的当前最大 ID 及末尾几行数据
      供配表流程确认推断 ID 时使用

用法：
    python read_sheet_info.py --sheet PicPanel
    python read_sheet_info.py --sheet PicPanel --tail 5
    python read_sheet_info.py --list-sheets

依赖：openpyxl（只读，不修改文件）
"""

import argparse
import sys
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: 缺少 openpyxl，请先安装：pip install openpyxl")
    sys.exit(1)

FILE_PATH = r"F:\OL2wc\NBA2KOL2Doc_proj\PlayerSkin.xlsx"


def check_lock(filepath):
    lock = os.path.join(os.path.dirname(filepath), "~$" + os.path.basename(filepath))
    if os.path.exists(lock):
        print(f"警告：检测到锁定文件 {lock}，文件可能正在被 Excel 打开（只读模式依然可以读取）")


def load_workbook(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: 文件不存在：{filepath}")
        sys.exit(1)
    return openpyxl.load_workbook(filepath, read_only=True, data_only=True)


def list_sheets(wb):
    print("可用 sheet 列表：")
    for name in wb.sheetnames:
        print(f"  - {name}")


def get_sheet_info(wb, sheet_name, tail=3):
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet '{sheet_name}' 不存在。可用：{wb.sheetnames}")
        sys.exit(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # 找数据起始行（跳过表头，通常第1~4行为表头，第5行起为数据）
    # 用非空 ID 来判断
    data_rows = []
    header_rows = []
    for i, row in enumerate(rows):
        if row[0] is None:
            header_rows.append((i + 1, row))
        else:
            # 尝试判断是否为数字 ID（跳过字符串表头）
            try:
                int(row[0])
                data_rows.append((i + 1, row))
            except (ValueError, TypeError):
                header_rows.append((i + 1, row))

    if not data_rows:
        print(f"sheet '{sheet_name}' 中暂无数据行（ID 列为空或无法解析为数字）")
        return

    max_id_row = max(data_rows, key=lambda x: int(x[1][0]))
    max_id = int(max_id_row[1][0])

    print(f"\n=== {sheet_name} ===")
    print(f"当前最大 ID: {max_id}")
    print(f"推断下一行 ID: {max_id + 1}")
    print(f"\n末尾 {min(tail, len(data_rows))} 行数据（行号, 前5列）：")
    for row_num, row in data_rows[-tail:]:
        preview = [str(v) if v is not None else "" for v in row[:5]]
        print(f"  行{row_num}: {preview}")


def main():
    parser = argparse.ArgumentParser(description="读取 PlayerSkin.xlsx sheet 最大 ID 信息")
    parser.add_argument("--sheet", "-s", help="sheet 名称，如 PicPanel")
    parser.add_argument("--tail", "-t", type=int, default=3, help="显示末尾几行，默认 3")
    parser.add_argument("--list-sheets", "-l", action="store_true", help="列出所有 sheet 名")
    parser.add_argument("--excel-dir", default=None, help="Excel 文件目录（默认使用内置路径）")
    args = parser.parse_args()

    filepath = FILE_PATH
    if args.excel_dir:
        filepath = os.path.join(args.excel_dir, "PlayerSkin.xlsx")

    check_lock(filepath)
    wb = load_workbook(filepath)

    if args.list_sheets:
        list_sheets(wb)
        return

    if not args.sheet:
        print("请指定 --sheet 参数，或用 --list-sheets 查看可用 sheet")
        parser.print_help()
        sys.exit(1)

    get_sheet_info(wb, args.sheet, args.tail)
    wb.close()


if __name__ == "__main__":
    main()
